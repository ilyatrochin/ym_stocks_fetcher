"""
fetch_ym_report.py — модуль 2.5: исторические данные через Reports API.

Зачем:
  Endpoint POST /v1/businesses/{id}/orders молча игнорирует dateFrom/dateTo
  и возвращает только последние ~30 дней. Чтобы получить историю
  заказов/реализации старше 30 дней, нужен асинхронный отчётный API.

Что делает:
  1. POST /reports/goods-realization/generate с businessId/year/month
     (или dateFrom/dateTo — поддерживаются оба формата) → получает reportId.
  2. Поллит GET /reports/info/{reportId} раз в 10 сек пока status=DONE.
  3. Скачивает XLSX-файл по url, сохраняет на диск.
  4. (Опц.) Парсит XLSX в JSON и печатает структуру (диагностика колонок).

  Запись в Google Sheets делается отдельным шагом — сначала надо понять
  структуру колонок, она зависит от типа отчёта.

Запуск (диагностика):
  # Скачать отчёт за апрель 2026 и показать структуру
  python fetch_ym_report.py --year 2026 --month 4 --inspect

  # Только сгенерировать (вернёт reportId)
  python fetch_ym_report.py --year 2026 --month 4 --generate-only

  # Скачать уже готовый отчёт по reportId
  python fetch_ym_report.py --report-id <id> --download-only

Запуск (нормальный):
  python fetch_ym_report.py --year 2026 --month 1
  # Файл уйдёт в /tmp/ym_reports/goods_realization_2026-01.xlsx
"""
from __future__ import annotations
import argparse
import json
import logging
import os
import sys
import time
from datetime import date, datetime
from pathlib import Path

import requests

# Локальные импорты (парсер XLSX + работа с Sheets)
try:
    from parse_realization import parse_realization_xlsx, aggregate_by_day_sku
except ImportError:
    parse_realization_xlsx = None  # type: ignore
    aggregate_by_day_sku = None  # type: ignore

API_BASE = "https://api.partner.market.yandex.ru"
GENERATE_ENDPOINT = "/reports/goods-realization/generate"
INFO_ENDPOINT = "/reports/info/{report_id}"

POLL_INTERVAL = 10        # сек между проверками статуса
POLL_TIMEOUT = 600        # 10 минут максимум на генерацию
HTTP_TIMEOUT = 60         # для скачивания файла
DEFAULT_OUTPUT_DIR = "/tmp/ym_reports"


class YMReportsClient:
    def __init__(self, api_key: str, business_id: int,
                 campaign_id: int | None = None):
        self.api_key = api_key
        self.business_id = business_id
        self.campaign_id = campaign_id
        self.session = requests.Session()
        self.session.headers.update({"Api-Key": api_key})

    def list_campaigns(self) -> list[dict]:
        """GET /v2/campaigns → список кампаний (магазинов) для текущего токена.

        Возвращает [{"id": ..., "domain": ..., "clientId": ...,
                     "business": {"id": ..., "name": ...}}, ...]
        """
        url = f"{API_BASE}/v2/campaigns"
        resp = self.session.get(url, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")
        data = resp.json()
        # Структура: {"campaigns": [...]} или {"result": {"campaigns": [...]}}
        return (data.get("campaigns")
                or (data.get("result") or {}).get("campaigns")
                or [])

    def generate(self, year: int | None = None, month: int | None = None,
                 date_from: str | None = None, date_to: str | None = None) -> str:
        """Запускает генерацию отчёта. Возвращает reportId.

        Goods-realization требует campaignId (привязан к магазину, не к бизнесу).
        Если campaign_id не задан явно — пробуем взять из env YM_CAMPAIGN_ID,
        иначе запросим список и используем первую кампанию.
        """
        url = f"{API_BASE}{GENERATE_ENDPOINT}"

        # ── Разрешаем campaignId ──────────────────────────────────────
        if not self.campaign_id:
            campaigns = self.list_campaigns()
            if not campaigns:
                raise RuntimeError("No campaigns found for this token. "
                                   "Token may lack campaign access.")
            if len(campaigns) > 1:
                names = [(c.get("id"), c.get("domain") or c.get("clientId"))
                         for c in campaigns]
                logging.warning(
                    "Multiple campaigns found, using first. "
                    "Set YM_CAMPAIGN_ID or --campaign-id to pick. "
                    "Available: %s", names)
            self.campaign_id = int(campaigns[0]["id"])
            logging.info("Resolved campaignId=%d from /v2/campaigns",
                         self.campaign_id)

        body: dict = {
            "businessId": self.business_id,
            "campaignId": self.campaign_id,
        }
        if year and month:
            body["year"] = year
            body["month"] = month
        elif date_from and date_to:
            body["dateFrom"] = date_from
            body["dateTo"] = date_to
        else:
            raise ValueError("Need either year+month or dateFrom+dateTo")

        # format=FILE → XLSX (по умолчанию), CSV/JSON тоже доступны
        params = {"format": "FILE"}

        logging.info("POST %s body=%s", GENERATE_ENDPOINT, body)
        resp = self.session.post(url, params=params, json=body,
                                 timeout=HTTP_TIMEOUT,
                                 headers={"Content-Type": "application/json"})
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")
        data = resp.json()
        result = data.get("result") or data
        report_id = result.get("reportId") or result.get("id")
        if not report_id:
            raise RuntimeError(f"No reportId in response: {data}")
        logging.info("Report generation started: reportId=%s", report_id)
        return str(report_id)

    def wait_ready(self, report_id: str,
                   poll_interval: int = POLL_INTERVAL,
                   timeout: int = POLL_TIMEOUT) -> dict:
        """Поллит статус отчёта пока status=DONE. Возвращает финальный result."""
        url = f"{API_BASE}{INFO_ENDPOINT.format(report_id=report_id)}"
        started = time.time()
        attempt = 0
        while True:
            attempt += 1
            resp = self.session.get(url, timeout=HTTP_TIMEOUT)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")
            data = resp.json()
            result = data.get("result") or data
            status = result.get("status", "")
            logging.info("[poll #%d] status=%s", attempt, status)

            if status == "DONE":
                return result
            if status in ("FAILED", "ERROR"):
                raise RuntimeError(f"Report failed: {result}")
            if time.time() - started > timeout:
                raise RuntimeError(
                    f"Timeout after {timeout}s, last status: {status}")
            time.sleep(poll_interval)

    def download(self, file_url: str, output_path: Path) -> Path:
        """Скачивает файл по URL из result.file."""
        logging.info("Downloading from %s", file_url)
        resp = self.session.get(file_url, timeout=HTTP_TIMEOUT, stream=True)
        if resp.status_code != 200:
            raise RuntimeError(f"Download failed: HTTP {resp.status_code}")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        size = output_path.stat().st_size
        logging.info("Saved %s (%d bytes)", output_path, size)
        return output_path


def inspect_xlsx(path: Path, sheet_filter: str | None = None,
                 max_rows: int = 25, full_width: bool = False) -> None:
    """Печатает структуру XLSX.

    sheet_filter — если задан, печатает ТОЛЬКО лист с этой подстрокой в имени
                   (регистронезависимо) и показывает больше строк.
    max_rows     — сколько строк показывать.
    full_width   — печатать значения без обрезки до 40 символов.

    NB: НЕ используем read_only=True — этот режим в openpyxl плохо работает
    с файлами без default style и часто врёт про max_row/max_column.
    """
    try:
        import openpyxl
    except ImportError:
        logging.error("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n=== {path.name} ===")
    print(f"Листы: {wb.sheetnames}\n")

    if sheet_filter:
        max_rows = max(max_rows, 60)

    def fmt(v):
        if v is None:
            return ""
        s = str(v)
        return s if full_width else s[:40]

    for sheet_name in wb.sheetnames:
        if sheet_filter and sheet_filter.lower() not in sheet_name.lower():
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        while all_rows and all(v is None or v == "" for v in all_rows[-1]):
            all_rows.pop()

        if not all_rows:
            print(f"--- Лист «{sheet_name}» (пусто) ---\n")
            continue

        n_rows = len(all_rows)
        max_col = 0
        for row in all_rows:
            non_empty_idx = [i for i, v in enumerate(row)
                             if v is not None and v != ""]
            if non_empty_idx:
                max_col = max(max_col, max(non_empty_idx) + 1)

        print(f"--- Лист «{sheet_name}» ({n_rows} строк × {max_col} колонок) ---")

        head_n = min(max_rows, n_rows)
        for row_idx, row in enumerate(all_rows[:head_n], start=1):
            values = [fmt(v) for v in row[:max_col]]
            print(f"  стр.{row_idx}: {values}")

        if n_rows > head_n:
            print(f"  ... ещё {n_rows - head_n} строк ...")
            mid = n_rows // 2
            for row_idx in [mid - 1, mid, mid + 1]:
                if 0 <= row_idx < n_rows:
                    row = all_rows[row_idx]
                    values = [fmt(v) for v in row[:max_col]]
                    print(f"  стр.{row_idx + 1} (середина): {values}")
            print(f"  ...")
            last_row = all_rows[-1]
            values = [fmt(v) for v in last_row[:max_col]]
            print(f"  стр.{n_rows} (последняя): {values}")
        print()
    wb.close()


def run(year: int | None, month: int | None,
        date_from: str | None, date_to: str | None,
        report_id: str | None, download_only: bool,
        generate_only: bool, inspect: bool, output_dir: str,
        campaign_id: int | None = None,
        list_campaigns: bool = False,
        import_and_download: bool = False,
        source_month: str | None = None,
        dry_run: bool = False) -> int:
    api_key = os.environ["YM_API_KEY"]
    business_id = int(os.environ["YM_BUSINESS_ID"])

    # campaign_id: явный аргумент → env → автодетект из /v2/campaigns
    if campaign_id is None:
        env_cid = os.environ.get("YM_CAMPAIGN_ID")
        if env_cid:
            campaign_id = int(env_cid)

    client = YMReportsClient(api_key, business_id, campaign_id=campaign_id)

    # Команда --list-campaigns: просто показать кампании и выйти
    if list_campaigns:
        campaigns = client.list_campaigns()
        if not campaigns:
            print("No campaigns found.")
            return 1
        print(f"\nНайдено кампаний: {len(campaigns)}\n")
        for c in campaigns:
            cid = c.get("id")
            domain = c.get("domain") or "(no domain)"
            client_id = c.get("clientId", "?")
            biz = c.get("business") or {}
            biz_name = biz.get("name", "")
            biz_id = biz.get("id", "")
            print(f"  campaignId = {cid}")
            print(f"    domain    : {domain}")
            print(f"    clientId  : {client_id}")
            print(f"    business  : {biz_name} (id={biz_id})")
            print()
        print("Чтобы использовать конкретную кампанию:")
        print(f"  export YM_CAMPAIGN_ID={campaigns[0].get('id')}")
        print("или передавайте --campaign-id <ID> аргументом.")
        return 0

    out_dir = Path(output_dir)

    # 1. Получить reportId
    if not report_id:
        report_id = client.generate(year=year, month=month,
                                    date_from=date_from, date_to=date_to)
        if generate_only:
            print(f"reportId={report_id}")
            print(f"Чтобы скачать когда будет готов:")
            print(f"  python fetch_ym_report.py --report-id {report_id} "
                  f"--download-only")
            return 0

    # 2. Дождаться готовности
    result = client.wait_ready(report_id)
    file_url = result.get("file") or result.get("url")
    if not file_url:
        logging.error("DONE без файла? Ответ: %s", result)
        return 1

    # 3. Скачать
    suffix = ".xlsx"
    if year and month:
        fname = f"goods_realization_{year:04d}-{month:02d}{suffix}"
        sm = source_month or f"{year:04d}-{month:02d}"
    elif date_from and date_to:
        fname = f"goods_realization_{date_from}_{date_to}{suffix}"
        sm = source_month
    else:
        fname = f"goods_realization_{report_id}{suffix}"
        sm = source_month
    output_path = out_dir / fname
    client.download(file_url, output_path)

    # 4. (Опц.) Диагностика структуры
    if inspect:
        inspect_xlsx(output_path)

    # 5. (Опц.) Сразу импортировать в sales_history
    if import_and_download:
        logging.info("Importing %s into sales_history...", output_path.name)
        res = import_to_sheets(output_path, source_month=sm, dry_run=dry_run)
        print(f"📥 Импорт: {res['parsed_rows']} строк → "
              f"{res['aggregated_rows']} агрегатов → "
              f"{res['written_rows']} записано в sales_history")

    print(f"\n✅ Готово: {output_path}")
    return 0


def import_to_sheets(xlsx_path: Path, source_month: str | None = None,
                     dry_run: bool = False) -> dict:
    """Парсит XLSX и записывает агрегаты в лист sales_history Google Sheets.

    Структура листа sales_history:
       date | sku | qty_delivered | source_month
    Поведение upsert: удаляет старые строки с тем же source_month, пишет свежие.
    Это позволяет безопасно перезаливать тот же месяц.

    Возвращает {parsed_rows, aggregated_rows, written_rows, skipped}.
    """
    if parse_realization_xlsx is None:
        raise RuntimeError("parse_realization.py not found in PYTHONPATH")

    rows = parse_realization_xlsx(xlsx_path, source_month=source_month)
    agg = aggregate_by_day_sku(rows)
    result = {
        "parsed_rows": len(rows),
        "aggregated_rows": len(agg),
        "written_rows": 0,
        "skipped": 0,
    }

    if dry_run:
        logging.info("[dry-run] Would write %d rows to sales_history", len(agg))
        # Печатаем первые 10 для проверки
        print("\nПервые 10 строк агрегата:")
        for d, sku, qty, src in agg[:10]:
            print(f"  {d.isoformat()}  {sku:<6} qty={qty}  [{src}]")
        if len(agg) > 10:
            print(f"  ... и ещё {len(agg) - 10}")
        return result

    # Запись в Sheets
    try:
        import gspread
        from oauth2client.service_account import ServiceAccountCredentials
    except ImportError as e:
        raise RuntimeError("gspread/oauth2client not installed") from e

    sheet_id = os.environ["GSHEET_ID"]
    creds_path = os.environ.get("GOOGLE_CREDS",
                                "/opt/openclaw/secrets/gcp-sa.json")
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
    spreadsheet = gspread.authorize(creds).open_by_key(sheet_id)

    SHEET_NAME = "sales_history"
    HEADERS = ["date", "sku", "qty_delivered", "source_month"]

    # Создаём лист если его нет
    try:
        ws = spreadsheet.worksheet(SHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        logging.info("Creating new sheet '%s'", SHEET_NAME)
        ws = spreadsheet.add_worksheet(title=SHEET_NAME, rows=10000, cols=4)
        ws.update("A1:D1", [HEADERS])

    # Находим строки текущего source_month и удаляем
    all_values = ws.get_all_values()
    if all_values and all_values[0][:4] != HEADERS:
        # Лист есть, но заголовок не совпадает — записываем
        ws.update("A1:D1", [HEADERS])

    source_months_in_agg = {src for _, _, _, src in agg}
    rows_to_delete: list[int] = []
    for i, row in enumerate(all_values[1:], start=2):  # 1-indexed для gspread
        if len(row) >= 4 and row[3] in source_months_in_agg:
            rows_to_delete.append(i)

    if rows_to_delete:
        logging.info("Deleting %d existing rows for source_months %s",
                     len(rows_to_delete), source_months_in_agg)
        # Удаляем с конца чтобы индексы не съезжали. gspread.delete_rows
        # принимает диапазоны — но проще батчами с конца
        # Группируем подряд идущие
        rows_to_delete.sort(reverse=True)
        i = 0
        while i < len(rows_to_delete):
            end = rows_to_delete[i]
            start = end
            while (i + 1 < len(rows_to_delete)
                   and rows_to_delete[i + 1] == start - 1):
                start = rows_to_delete[i + 1]
                i += 1
            ws.delete_rows(start, end)
            i += 1

    # Готовим строки на запись
    rows_to_write = [[d.isoformat(), sku, qty, src]
                     for d, sku, qty, src in agg]

    if rows_to_write:
        ws.append_rows(rows_to_write, value_input_option="USER_ENTERED")
        result["written_rows"] = len(rows_to_write)
        logging.info("Wrote %d rows to %s", len(rows_to_write), SHEET_NAME)

    return result


def main():
    p = argparse.ArgumentParser(
        description="Отчётный API YM (исторические данные)")
    p.add_argument("--year", type=int, help="Год отчёта")
    p.add_argument("--month", type=int, help="Месяц отчёта (1-12)")
    p.add_argument("--from", dest="date_from", help="Альтернатива year/month: "
                                                    "начало периода ISO")
    p.add_argument("--to", dest="date_to", help="Альтернатива year/month: "
                                                "конец периода ISO")
    p.add_argument("--report-id", help="Использовать готовый reportId "
                                       "вместо генерации нового")
    p.add_argument("--campaign-id", type=int, default=None,
                   help="ID кампании (магазина). Если не задан, берётся из "
                        "env YM_CAMPAIGN_ID, иначе автодетект из /v2/campaigns "
                        "(первая в списке).")
    p.add_argument("--list-campaigns", action="store_true",
                   help="Показать список доступных кампаний и выйти")
    p.add_argument("--generate-only", action="store_true",
                   help="Только запустить генерацию, не ждать готовности")
    p.add_argument("--download-only", action="store_true",
                   help="С --report-id: пропустить генерацию, "
                        "сразу ждать и скачивать")
    p.add_argument("--inspect", action="store_true",
                   help="После скачивания распечатать структуру XLSX")
    p.add_argument("--inspect-file", metavar="PATH",
                   help="Только распечатать структуру уже скачанного XLSX "
                        "(API не вызывается)")
    p.add_argument("--inspect-sheet", metavar="NAME",
                   help="При inspect показывать только лист с этим именем "
                        "(подстрока, регистронезависимо). "
                        "Например: --inspect-sheet 'Доставленные'")
    p.add_argument("--full-width", action="store_true",
                   help="Не обрезать значения ячеек до 40 символов")
    p.add_argument("--import", dest="import_xlsx", metavar="PATH",
                   help="Распарсить XLSX и залить в лист sales_history "
                        "в Google Sheets (API не вызывается)")
    p.add_argument("--import-and-download", action="store_true",
                   help="После скачивания отчёта сразу импортировать "
                        "его в sales_history")
    p.add_argument("--source-month", metavar="YYYY-MM",
                   help="Метка месяца для sales_history. Если не задана, "
                        "берётся из имени файла.")
    p.add_argument("--dry-run", action="store_true",
                   help="При импорте — не писать в Sheets, только показать "
                        "что будет записано")
    p.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                   help=f"Куда сохранять (по умолчанию {DEFAULT_OUTPUT_DIR})")
    p.add_argument("--verbose", "-v", action="store_true")
    args = p.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s")

    # --inspect-file: автономный режим, без API
    if args.inspect_file:
        inspect_xlsx(Path(args.inspect_file),
                     sheet_filter=args.inspect_sheet,
                     full_width=args.full_width)
        sys.exit(0)

    # --import: автономный импорт уже скачанного файла, без API
    if args.import_xlsx:
        try:
            res = import_to_sheets(Path(args.import_xlsx),
                                   source_month=args.source_month,
                                   dry_run=args.dry_run)
            print(f"\n✅ {res['parsed_rows']} строк распарсено, "
                  f"{res['aggregated_rows']} агрегатов, "
                  f"{res['written_rows']} записано в sales_history")
            sys.exit(0)
        except Exception as e:
            logging.exception("Import failed: %s", e)
            sys.exit(1)

    if not args.list_campaigns and not args.report_id and \
       not (args.year and args.month) and \
       not (args.date_from and args.date_to):
        p.error("Need either --year+--month, --from+--to, --report-id, "
                "--list-campaigns, --inspect-file, or --import")

    try:
        rc = run(
            year=args.year, month=args.month,
            date_from=args.date_from, date_to=args.date_to,
            report_id=args.report_id,
            download_only=args.download_only,
            generate_only=args.generate_only,
            inspect=args.inspect,
            output_dir=args.output_dir,
            campaign_id=args.campaign_id,
            list_campaigns=args.list_campaigns,
            import_and_download=args.import_and_download,
            source_month=args.source_month,
            dry_run=args.dry_run,
        )
        sys.exit(rc)
    except KeyError as e:
        logging.error("Missing env var: %s", e)
        sys.exit(1)
    except Exception as e:
        logging.exception("Failed: %s", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
