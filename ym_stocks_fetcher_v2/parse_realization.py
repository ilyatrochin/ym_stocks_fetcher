"""
parse_realization.py — парсер XLSX-отчёта YM «Доставленные товары».

Структура листа (для отчёта goods-realization):
- Строки 1-16: реквизиты договора, заказчик, исполнитель
- Строка 17: заголовок таблицы
- Строки 18..N-1: данные
- Строка N (последняя): "Итого:"

Колонки (нумерация с 0):
  0  Номер заказа              ← orderId
  1  Ваш номер заказа
  2  Тип заказа                ← "Продажа физлицу" / "Возврат от физлица"
  3  Название товара
  4  Ваш SKU                   ← offerId (то же что в API)
  5  SKU на складе
  6  Количество переданных в доставку, шт.
  7  Доставлено, шт.           ← qty
  8  Дата оформления заказа    ← DD.MM.YYYY
  9  Дата передачи товара в доставку
 10  Дата доставки товара      ← DD.MM.YYYY
 ...

Возвращает агрегаты по (date, sku) → qty_delivered.
"""
from __future__ import annotations
import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path


@dataclass
class RealizationRow:
    date_delivered: date
    sku: str
    qty: int
    source_month: str  # "2026-04"


# Колонки в листе «Доставленные товары» (0-indexed)
COL_ORDER_TYPE = 2
COL_SKU = 4
COL_QTY_DELIVERED = 7
COL_DATE_DELIVERED = 10

# Какие типы заказов считаем продажами
SALE_TYPES = {"Продажа физлицу", "Продажа юрлицу"}

# Где-то на этих строках обычно лежит заголовок таблицы.
# Ищем строку, где первая ячейка == "Номер заказа".
HEADER_MARKER = "Номер заказа"


def _parse_date(value) -> date | None:
    """Принимает datetime, date или строку 'DD.MM.YYYY'/'YYYY-MM-DD'."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_realization_xlsx(path: Path, source_month: str | None = None,
                           sheet_name: str = "Доставленные товары"
                           ) -> list[RealizationRow]:
    """Парсит лист «Доставленные товары» из XLSX, возвращает строки продаж.

    source_month — для трассировки источника, например "2026-04".
                   Если не задано, парсится из имени файла.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl not installed; run: "
                           "pip install openpyxl") from e

    path = Path(path)
    if source_month is None:
        # goods_realization_2026-04.xlsx → "2026-04"
        stem = path.stem
        if "_" in stem:
            source_month = stem.rsplit("_", 1)[-1]
        else:
            source_month = stem

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found in {path.name}. "
            f"Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Находим строку заголовка
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and str(row[0]).strip() == HEADER_MARKER:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(
            f"Header row 'Номер заказа' not found in {path.name}")

    logging.info("Header found at row %d (1-indexed: %d)",
                 header_row_idx, header_row_idx + 1)

    # Парсим данные
    parsed: list[RealizationRow] = []
    skipped_returns = 0
    skipped_no_date = 0
    skipped_no_sku = 0
    skipped_total = 0  # строка "Итого:"

    for row in rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue

        first = str(row[0]).strip()
        if first.startswith("Итого"):
            skipped_total += 1
            continue

        order_type = str(row[COL_ORDER_TYPE] or "").strip()
        if order_type not in SALE_TYPES:
            # Возвраты не учитываем — они уже в "Возвращенные товары",
            # к тому же занижают скорость продаж артифициально.
            skipped_returns += 1
            continue

        sku = str(row[COL_SKU] or "").strip()
        if not sku:
            skipped_no_sku += 1
            continue

        d = _parse_date(row[COL_DATE_DELIVERED])
        if d is None:
            skipped_no_date += 1
            continue

        try:
            qty = int(float(row[COL_QTY_DELIVERED] or 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue

        parsed.append(RealizationRow(
            date_delivered=d,
            sku=sku,
            qty=qty,
            source_month=source_month,
        ))

    logging.info(
        "Parsed %s: %d rows | skipped returns=%d, no_date=%d, no_sku=%d, "
        "total=%d", path.name, len(parsed),
        skipped_returns, skipped_no_date, skipped_no_sku, skipped_total)

    return parsed


def aggregate_by_day_sku(rows: list[RealizationRow]
                         ) -> list[tuple[date, str, int, str]]:
    """Сворачивает построчные данные в (date, sku, qty, source_month).

    Несколько строк на тот же день+SKU суммируются.
    """
    bucket: dict[tuple[date, str, str], int] = defaultdict(int)
    for r in rows:
        bucket[(r.date_delivered, r.sku, r.source_month)] += r.qty

    out = sorted(
        [(d, sku, qty, src) for (d, sku, src), qty in bucket.items()],
        key=lambda x: (x[0], x[1]),
    )
    return out


if __name__ == "__main__":
    # CLI для отладки: python parse_realization.py path/to/file.xlsx
    import sys
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")
    if len(sys.argv) < 2:
        print("Usage: python parse_realization.py <path.xlsx>")
        sys.exit(1)
    rows = parse_realization_xlsx(Path(sys.argv[1]))
    print(f"\nDetailed: {len(rows)} sale lines")

    agg = aggregate_by_day_sku(rows)
    print(f"Aggregated (day × sku): {len(agg)} groups\n")

    # Сводка по SKU
    by_sku = defaultdict(int)
    for r in rows:
        by_sku[r.sku] += r.qty
    print("По SKU:")
    for sku, q in sorted(by_sku.items(), key=lambda x: -x[1]):
        print(f"  {sku}: {q} шт")

    # Период
    if rows:
        dates = [r.date_delivered for r in rows]
        n_days = (max(dates) - min(dates)).days + 1
        print(f"\nПериод: {min(dates)} .. {max(dates)} ({n_days} дней)")
        print(f"Всего продаж: {sum(r.qty for r in rows)} шт")
        print(f"Средняя скорость по всем SKU: "
              f"{sum(r.qty for r in rows) / n_days:.2f} шт/день")
